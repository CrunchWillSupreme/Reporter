"""
LeX Examiner Updates
Get new examiners
"""
import pandas as pd, datetime, pyodbc, win32com.client
from openpyxl import load_workbook

def main():
    today = datetime.datetime.today()
    path = r'\PATH\TO\FILE\ExaminerUpdateFormula.xlsm'
    macro = 'MUExaminer.MUExaminer'
    saveas = r"\\PATH\TO\FILE\UnMapped_Examiners_{today:%m_%d_%y}.xlsx".format(today = today)
    sql = read_sql()
    data = run_query(sql)
    save_excel(data)
    call_macro(wb_path = path, macro_name = macro, save_path = saveas)

def read_sql():
    Today = datetime.datetime.today()
    last_month = Today - datetime.timedelta(days=30)
    last_run = last_month.strftime('%Y-%m-%d')
    print('Reading ExaminerUpdate.sql file...')
    with open(r'\\PATH\TO\FILE\ExaminerUpdate.sql') as f:
        sql = f.read().format(last_run=last_run)
        print('SQL file read.')
    return sql

def run_query(sql):
    cnn = pyodbc.connect('DRIVER={SQL Server};PORT=1433;SERVER=[SERVER_ADDRESS]')
    print('Running SQL query...')
    data = pd.read_sql(sql, cnn)
    print('SQL query finished running')
    return data
    
def save_excel(data):
    print('Beginning savewb method...')
    rngoutput = data.values.tolist()
    print('Sheet added!\nRetrieving TEMPLATE workbook')
    wb = load_workbook(r"\\PATH\TO\FILE\ExaminerUpdateFormulaTemplate.xlsm", keep_vba=True)
    print("TEMPLATE workbook retrieved! \nPasting df output onto Excel sheet..")
    ws=wb.get_sheet_by_name('Sheet1')
    for row_num, row in enumerate(rngoutput):
        for col_num,val in enumerate(row):
            ws.cell(row=row_num+2,column=col_num+1).value=val #python is zero-indexed, openpyxl is 1-indexed
    print("output pasted onto sheet!\nSaving workbook...")                    
    wb.save(r"\\PATH\TO\FILE\ExaminerUpdateFormula.xlsm")
    print('Workbook Saved!')

def call_macro(wb_path, macro_name,*args, save_path = None):
    """
    Opens an excel workbook wb_path, calls its macro_name method with
    parameters *args. Optionally saves to save_path.

    Parameters:
    -----
    wb_path str:
    path to excel xlsm workbook containing macro

    macro_name str:
    module and function/subroutine to invoke

    *args tuple:
    tuple of positional arguments to pass to the excel macro

    save_path str:
    path and name to save workbook as after running the macro
    """
    #grab excel application
    xl = win32com.client.Dispatch("Excel.Application")
    print(f'loaded excel as "{xl}".')
    #grab actual workbook
    wb = xl.Workbooks.Open(Filename=wb_path)
    
    print(f'loaded workbook as "{wb.Name}"')
    #call macro from workbook
    print(f'calling {wb.Name}!{macro_name}...',*args)
    if any(args):
        xl.Application.Run(f'{wb.Name}!{macro_name}',*args)
    else:
        xl.Application.Run(f'{wb.Name}!{macro_name}')
        
    #save at save_path, or over the original if save_path not specified
    print('saving')
    if save_path is not None:
         xl.Application.Run(f'{wb.Name}!Savexlsx', save_path)
#         subprocess.call([r"PATH\TO\FILE\AutoHotkey.exe", r"\\PATH\TO\FILE\Enter.ahk"])
#         process = subprocess.Popen([r"PATH\TO\FILE\AutoHotkey.exe",r"\\PATH\TO\FILE\Enter.ahk"])
#         process.wait()
    #cleanup
    print('quitting')
    xl.Application.DisplayAlerts = False
    xl.Application.Quit()  
    
main()
