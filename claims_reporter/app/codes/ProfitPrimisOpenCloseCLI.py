""" 
			Profit and Primis Open and Closed Claims CLI
			10/11/18
			Created by Will Han
"""

import pyodbc, pandas as pd, datetime as dt, os, json, smtplib, argparse, sys, time
from openpyxl import load_workbook, Workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from smtplib import SMTPSenderRefused

def main(email_user, email_pwd, datestring, sendmail:bool=False):
    schedule = get_json()
    d_string = date_format(datestring)
    start_date, end_date = get_open_and_closed_dates(schedule, dates_for=d_string)
#    tempquery = tempQuery()
#    finalquery = finalQuery(start_date, end_date)
    source = 'Profit and Primis'
    #	temp = tempquery
    #	final = finalquery
    temp, finalquery = read_queries(start_date, end_date)
    runqueryandexcel(source, temp, finalquery, end_date)
    if sendmail:
        flag = False
        while not flag:
            try:    
                recipients = ['USER@EMAIL.com']
                subject = "Profit & Primis Open and Closed Claim Report"
                body = "Hello USER,\n\nAttached you will find the Profit & Primis Opened and Closed Claims Report for {end_date:%B %Y}.  Please let me know if you have any questions.\n\nThanks,\nWill Han".format(end_date=end_date)
                ppattach = [r'\\PATH\TO\FILE\Profit & Primis Opened and Closed Claim Report - {end_date:%B %Y}.xlsx'.format(end_date=end_date)]
                email( recipients, subject, body, ppattach, email_user, email_pwd)
                print('Email Sent!')
                flag = True
            except SMTPSenderRefused:
                time.sleep(5)
	
def get_json():
	with open(r'\\PATH\TO\FILE\schedule.json') as f:
		schedule = json.loads(f.read())
	return schedule
	
def get_open_and_closed_dates(schedule, dates_for=dt.date.today()):
    month = str(dates_for.month)
    year = str(dates_for.year)
    start_date = dt.datetime.strptime(schedule[year][month]['start_date'],'%m/%d/%Y')
    end_date = dt.datetime.strptime(schedule[year][month]['end_date'],'%m/%d/%Y')
    return start_date, end_date

def read_queries(start_date, end_date):
    with open(r'PATH\TO\FILE\results_temp.sql') as r:
        temp = r.read()
    with open(r'PATH\TO\FILE\final_query.sql') as f:
        finalquery = f.read().format(start_date=start_date, end_date=end_date)
    return temp, finalquery
	
def runqueryandexcel(source, temp, finalquery, end_date):
    #Create Connection
    print('Creating server connection...')
    DataLakeserver = '[SERVER_ADDRESS]'
    driver = '{SQL Server}'    # Driver you need to connect to the database
    port = '1433'
    DataLakecnn = pyodbc.connect('DRIVER='+driver+';PORT='+port+';SERVER='+DataLakeserver)
    DataLakecnn.autocommit=True
    cursor=DataLakecnn.cursor()
    print("Executing "+source+" query...")
    cursor.execute(temp)
    print("Storing "+source+" query results as pd.DF...")
    output = pd.read_sql_query(finalquery, DataLakecnn)
    print(source + " results stored!")                               
    rngoutput = output.values.tolist()
    print('Sheet added!\nRetrieving TEMPLATE workbook')
    wb = load_workbook(r"\\PATH\TO\FILE\TEMPLATE - ProfitPrimisOpenedClosedClaim.xlsx")
    print("TEMPLATE workbook retrieved! \nPasting "+source+" output onto Excel sheet..")
    ws=wb.get_sheet_by_name('Sheet1')
    for row_num, row in enumerate(rngoutput):
        for col_num,val in enumerate(row):
            ws.cell(row=row_num+2,column=col_num+1).value=val #python is zero-indexed, openpyxl is 1-indexed
    print(source+" output pasted onto sheet!")                    
    print('Sheet name updated! \nSaving workbook...')
    wb.save(r'\\PATH\TO\FILE\Profit & Primis Opened and Closed Claim Report - {end_date:%B %Y}.xlsx'.format(end_date=end_date))
    print('Workbook Saved!')
    return

def email(to, subject, text, ppattach, email_user, email_pwd, cc=None):
    print('Assigning Sender, Recipient(s), and Subject of email...')
    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = ", ".join(to)
    bcc = ['USER@EMAIL.com']
    msg['Subject'] = subject
    print('Adding body of message...')
    msg.attach(MIMEText(text))
    print('Formatting attachments...')
    for file in ppattach:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(open(file, 'rb').read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(file))
        msg.attach(part)
        print('Setting up server...')   
        mailServer = smtplib.SMTP('outlook.markelcorp.com', 587)
        mailServer.ehlo()
        mailServer.starttls()
        mailServer.ehlo()
        mailServer.login(email_user, email_pwd)
        mailServer.sendmail(email_user, to+bcc, msg.as_string())
        # Should be mailServer.quit(), but that crashes...
        mailServer.close()
        print('Email Sent!')
      
def date_format(datestring):
    x = datestring
    x = dt.datetime.strptime(x, '%m/%Y')
    return x



if __name__ == '__main__':
    parser = argparse.ArgumentParser(sys.argv)
    parser.add_argument('-e', type = str, help = 'your outlook email address')
    parser.add_argument('-p', type = str, help = 'your outlook email password (same as your VDI password)')
    parser.add_argument('-d', type = str, help = 'start date (month/year)')
    parser.add_argument('--creds', type = str, help = 'the path to your credentials file')
    args = parser.args()
    if (args.e and args.p and args.d):
        main(args.e, args.p, args.d)
    elif (args.creds):
        with open(args.creds) as c:
            creds = json.loads(c.read())
        email = creds['c1']
        pw = creds['c2']
        main(email, pw)
    else:
        raise Exception('Input either your email address and password OR the path to your config file')
	
	
	
	
	
	
