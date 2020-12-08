
#ACKNOWLEDGEMENT LETTERS
import argparse
import pyodbc
import sys
import pandas as pd
#import xlsxwriter as xw
import datetime
import os
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
#from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook, Workbook
import ctypes
#from openpyxl.styles import Font

#def getcreds(cred_file=r'P:\PersonalConfig\credentials.json'):
#    with open(cred_file) as c:
#        creds = json.loads(c.read())
#        return creds

# from AckLetter import main as send_letters
# send_letters(username, password)

def main(username, password, sendmail=False):
    #Set Directory
    theday = datetime.date.today()
    print(f"Setting directory to folder: \PATH\TO\INPUT\{theday.year}\{datetime.datetime.strftime(theday,'%B')}")
    os.chdir(f"//PATH/TO/INPUT/{theday.year}/{datetime.datetime.strftime(theday,'%B')}/")
    #Create variable for date
    try:
        Primisclaims = cleanclaimnumber('Primis')
        Primisfinalquery = """SELECT * FROM ##TEMP_ADD_PRIMIS"""
        Primisquery = getPrimisquery(Primisclaims)
        runqueryandexcel('Primis',Primisquery,Primisfinalquery)
    except FileNotFoundError:
        print("No Primis claims for today!")
        
    try:
        ICONclaims = cleanclaimnumber('ICON')
        ICONfinalquery = """SELECT * FROM ##TEMP_ADD_ICON"""
        ICONquery = getICONquery(ICONclaims)
        runqueryandexcel('ICON',ICONquery,ICONfinalquery)
    except FileNotFoundError:
        print("No ICON claims for today!")
    if sendmail:
#        recipients = ['USER@COMPANY.com',]
#        cc = ['USER@COMPANY.com','USER@COMPANY.com', 'USER@COMPANY.com']
        recipients = ['USER@COMPANY.com']
        cc = ['USER@COMPANY.com','USER@COMPANY.com']
        subject = "Primis and ICON data request for Acknowledgement letters"
        body = "Hi USER,\n\nAttached you will find the ICON and PRIMIS extracts for the Acknowledgement letters.  Please let me know if you have any questions.\n\nThanks,\nWill Han"
        #Set up crap for the attachments
        #FOR ICON AND PRIMIS
        files = [r'\\PATH\TO\FILE\{0:%Y}\{0:%B}\{1} ICON Ack Letter Extract.xlsx'.format((theday),gettoday(False)),
                            r'\\PATH\TO\FILE\{0:%Y}\{0:%B}\{1} Primis Ack Letter Extract.xlsx'.format((theday),gettoday(False))]
        
        mail(recipients, subject, body, files, username, password, cc)
    Mbox("Report Finished!","Your report has finished running!") 
    
def gettoday(one=True):  #if we want today2, pass "False"
#    print('Creating date variable...')
    d= datetime.date.today()
    mon = d.month
    day = d.day
    year = d.year
    today = str(mon) + "-" + str(day)
    listy = [str(mon), str(day), str(year)]
    today2 = ".".join(listy)
    if one:
        return today
    return today2


def create_folder():
    newinputpath= f"//PATH/TO/INPUT/{datetime.date.today().year}/{datetime.datetime.strftime(datetime.date.today(),'%B')}/"
    if not os.path.exists(newinputpath):
        os.makedirs(newinputpath)
    newreportpath= f"//PATH/TO/REPORT/{datetime.date.today().year}/{datetime.datetime.strftime(datetime.date.today(),'%B')}/"
    if not os.path.exists(newreportpath):
        os.makedirs(newreportpath)
        
#This function does the above, commented out code
def cleanclaimnumber(source):
    print("Reading in "+source+ " excel file...")
    pdExcel = pd.ExcelFile(source +" "+ gettoday() + ".xlsx")
    print("Parse " +source+ " data to df...")
    parsed = pdExcel.parse(source, header = None)
    print("Adding quotes and commas to " +source+ " claim numbers...")
    listed = parsed.values.tolist()
    listed = [x[0].strip(' ') for x in listed]
    claims = "'" + "','".join(listed) + "'"
    print(source+ " claim numbers formatted!")
    return claims
#source='ICON'
#path = f"//PATH/TO/INPUT/{datetime.date.today().year}/{datetime.datetime.strftime(datetime.date.today(),'%B')}/"+source+" "+gettoday()+".xlsx"
#path = os.path.realpath(path)
#os.startfile(path)


def getICONquery(ICONclaims):
    #Create ICON Query
    print('Creating ICON query...')
    ICONquery = """IF OBJECT_ID('tempdb..##TEMP_ADD_ICON') IS NOT NULL DROP TABLE ##TEMP_ADD_ICON
    Select distinct
    ISNULL(CC.B69_CLAIM_OCCUR,'') AS "CLAIM FOLDER ID"
    ,ISNULL(CC.A00_PNUM,'') AS "POLICY NUMBER"
    ,ISNULL(CONVERT(VARCHAR(10),CAST(CC.B70_LOSS_DATE AS DATE),101),'') AS "DATE_OF_LOSS" 
    ,ISNULL(CONVERT(VARCHAR(10),CAST(CC.D43_REPORTED_DATE as Date),101),'') as 'REPORTED DATE'
    ,ISNULL(D87_ACC_STATE,'') as 'ACCIDENT STATE'
    ,ISNULL(CONVERT(VARCHAR(10),CAST(CC.H53_NDATE AS DATE),101),'') AS 'ENTERED DATE'
    ,SANDBOX_CLAIMS_OPS.data.ProperCase(ISNULL(NAD2.B27_NAME1,'')) AS "NAMED INSURED"
    ,ISNULL(NAD2.EMAIL_ADDRESS,'') as 'INSURED EMAIL'
    ,ISNULL(NAD2.B28_ADDR1,'')+ ' '+ ISNULL(NAD2.B28_ADDR2,'') as 'INSURED ADDRESS'
    ,ISNULL(NAD2.B30_CITY,'') as 'INSURED CITY'
    ,ISNULL(NAD2.B31_STATE,'') as 'INSURED STATE'
    ,ISNULL(NAD2.B32_ZIP,'') as 'INSURED ZIP'
    ,SANDBOX_CLAIMS_OPS.data.ProperCase(LTRIM(RTRIM(REPLACE(ISNULL(NA.B27_NAME1,''),'BI','')))) AS "CLAIMANT NAME"
    ,SANDBOX_CLAIMS_OPS.data.ProperCase(ISNULL(NAD.B27_NAME1,'')) AS "PRODUCER NAME"
    ,ISNULL(NAD.B28_ADDR1,'') as 'PRODUCER ADDRESS'
    ,ISNULL(NAD.B30_CITY,'') as 'PRODUCER CITY'
    ,ISNULL(NAD.B31_STATE,'') as 'PRODUCER STATE'
    ,ISNULL(NAD.B32_ZIP,'') as 'PRODUCER ZIP'
    ,ISNULL(NAD.EMAIL_ADDRESS,'') as 'PRODUCER EMAIL ADDRESS'
    ,ISNULL(ltrim(rtrim(replace(replace(NAD1.B27_NAME1,'(1st)',''),'(3rd)',''))),'') AS 'ADJUSTER NAME'
    ,ISNULL(CASE WHEN NADC.B27_NAME1 = 'Insurance Company of Evanston' THEN 'Evanston Insurance Company' ELSE NADC.B27_NAME1 End,'') as 'COMPANY NAME'
    --,ISNULL(LTRIM(RTRIM(REPLACE(DBO.Phone,'(Landline)',''))),'') as "Adjuster Phone"
    --,LTRIM(RTRIM(ISNULL(LEFT(DBO.[Email Address],CHARINDEX(' ',DBO.[Email Address])),NAD1.EMAIL_ADDRESS))) as 'EMAIL ADDRESS ADJUSTER'
    ,CASE WHEN NAD1.T50_WORKPHONE IS NULL THEN ''
        ELSE CONCAT('1+ (',LEFT(NAD1.T50_WORKPHONE,3),') ', SUBSTRING(NAD1.T50_WORKPHONE,4,3), '-', RIGHT(NAD1.T50_WORKPHONE,4)) END as "Adjuster Phone"
    ,LTRIM(RTRIM(ISNULL(NAD1.EMAIL_ADDRESS,''))) as 'EMAIL ADDRESS ADJUSTER'
    ,ISNULL(D74_CLAIM_TYPE,'') as 'Claim_Type'
    ,ISNULL(CC.R30_IN_HOUSE_ADJ,'') as 'Adjuster Code'
    
    INTO ##TEMP_ADD_ICON
    FROM RAW_ICON.ICON.CCOMMON CC
    
    LEFT JOIN RAW_ICON.ICON.CRESERVE CR
    ON CR.B69_CLAIM_OCCUR = CC.B69_CLAIM_OCCUR
    
    LEFT JOIN RAW_ICON.ICON.PCOMMON PC
    ON PC.A00_PNUM = CC.A00_PNUM
    AND PC.A06_EDITION = CC.A06_EDITION
    
    LEFT JOIN RAW_ICON.ICON.NAME_ADDRESS NA
    ON NA.E04_ORIGNUM COLLATE Latin1_General_CS_AS = CR.E04_CLAIMANT_NUM COLLATE Latin1_General_CS_AS
    AND NA.E04_NEXT IS NULL
    
    LEFT JOIN RAW_ICON.ICON.ACOMMON AC
    ON AC.A04_ANUM COLLATE Latin1_General_CS_AS= PC.A04_ANUM COLLATE Latin1_General_CS_AS
    
    LEFT JOIN RAW_ICON.ICON.NAME_ADDRESS NAD
    ON NAD.E04_ORIGNUM COLLATE Latin1_General_CS_AS  = AC.E04_ANAMNUM COLLATE Latin1_General_CS_AS 
    AND NAD.E04_NEXT IS NULL
    
    LEFT JOIN RAW_ICON.ICON.NAME_ADDRESS NAD1
    ON NAD1.E04_ORIGNUM COLLATE Latin1_General_CS_AS = CASE WHEN NAD1.B27_NAME1 = 'Record Only' THEN CC.R28_SUPERVISOR COLLATE Latin1_General_CS_AS ELSE CC.R30_IN_HOUSE_ADJ COLLATE Latin1_General_CS_AS END
    AND NAD1.E04_NEXT IS NULL
    
    LEFT JOIN RAW_ICON.ICON.NAME_ADDRESS NAD2
    ON NAD2.E04_ORIGNUM COLLATE Latin1_General_CS_AS = CC.E04_INSURED_NUM COLLATE Latin1_General_CS_AS
    AND NAD2.E04_NEXT IS NULL
    
    LEFT JOIN RAW_ICON.ICON.COMPANY CO
    ON CO.A01_COMPANY =CC.A01_COMPANY
    
    LEFT JOIN RAW_ICON.ICON.NAME_ADDRESS NADC
    ON NADC.E04_ORIGNUM = CO.E04_NAMENUM
    AND NADC.E04_NEXT IS NULL
    
    --LEFT JOIN SANDBOX_CLAIMS_OPS.data.['Directory by Organization'] DBO
    --ON DBO.[Worker] COLLATE SQL_Latin1_General_CP1_CI_AS= ISNULL(ltrim(rtrim(replace(replace(NAD1.B27_NAME1,'(1st)',''),'(3rd)',''))),'')
    
    WHERE AC.J03_PROD_NUM IS NULL
    and NAD1.B27_NAME1 NOT IN  ('To Be Assigned','Void')
    and CC.B69_CLAIM_OCCUR in (""" + ICONclaims + """)"""
    print('ICON query created!')
    return ICONquery

def getPrimisquery(Primisclaims):
    ##Create Primis Query
    Primisquery = """IF OBJECT_ID('tempdb..##TEMP_ADD_PRIMIS') IS NOT NULL DROP TABLE ##TEMP_ADD_PRIMIS
    Select distinct
    ISNULL(CC.B69_CLAIM_OCCUR,'') AS "CLAIM FOLDER ID"
    ,ISNULL(CC.A00_PNUM,'') AS "POLICY NUMBER"
    ,ISNULL(CONVERT(VARCHAR(10),CAST(CC.B70_LOSS_DATE AS DATE),101),'') AS "DATE_OF_LOSS" 
    ,ISNULL(CONVERT(VARCHAR(10),CAST(CC.D43_REPORTED_DATE as Date),101),'') as 'REPORTED DATE'
    ,ISNULL(D87_ACC_STATE,'') as 'ACCIDENT STATE'
    ,ISNULL(CONVERT(VARCHAR(10),CAST(CC.H53_NDATE AS DATE),101),'') AS 'ENTERED DATE'
    ,SANDBOX_CLAIMS_OPS.data.ProperCase(ISNULL(NAD2.B27_NAME1,'')) AS "NAMED INSURED"
    ,ISNULL(NAD2.E_MAIL_ADDRESS,'') as 'INSURED EMAIL'
    ,ISNULL(NAD2.B28_ADDR1,'')+ ' '+ ISNULL(NAD2.B28_ADDR2,'') as 'INSURED ADDRESS'
    ,ISNULL(NAD2.B30_CITY,'') as 'INSURED CITY'
    ,ISNULL(NAD2.B31_STATE,'') as 'INSURED STATE'
    ,ISNULL(NAD2.B32_ZIP,'') as 'INSURED ZIP'
    ,SANDBOX_CLAIMS_OPS.data.ProperCase(LTRIM(RTRIM(REPLACE(ISNULL(NA.B27_NAME1,''),'BI','')))) AS "CLAIMANT NAME"
    ,SANDBOX_CLAIMS_OPS.data.ProperCase(ISNULL(NAD.B27_NAME1,'')) AS "PRODUCER NAME"
    ,ISNULL(NAD.B28_ADDR1,'') as 'PRODUCER ADDRESS'
    ,ISNULL(NAD.B30_CITY,'') as 'PRODUCER CITY'
    ,ISNULL(NAD.B31_STATE,'') as 'PRODUCER STATE'
    ,ISNULL(NAD.B32_ZIP,'') as 'PRODUCER ZIP'
    ,ISNULL(NAD.E_MAIL_ADDRESS,'') as 'PRODUCER EMAIL ADDRESS'
    ,ISNULL(ltrim(rtrim(replace(replace(NAD1.B27_NAME1,'(1st)',''),'(3rd)',''))),'') AS 'ADJUSTER NAME'
    ,ISNULL(CASE WHEN NADC.B27_NAME1 = 'Insurance Company of Evanston' THEN 'Evanston Insurance Company' 
    			 WHEN NADC.B27_NAME1 = 'EIC (Formerly Essex)' THEN 'Evanston Insurance Company' ELSE NADC.B27_NAME1 End,'') as 'COMPANY NAME'
    --,ISNULL(LTRIM(RTRIM(REPLACE(DBO.Phone,'(Landline)',''))),'') as "Adjuster Phone"
    --,LTRIM(RTRIM(ISNULL(LEFT(DBO.[Email Address],CHARINDEX(' ',DBO.[Email Address])),NAD1.E_MAIL_ADDRESS))) as 'EMAIL ADDRESS ADJUSTER'
    ,CASE WHEN NAD1.T50_WORKPHONE IS NULL THEN ''
        ELSE CONCAT('1+ (',LEFT(NAD1.T50_WORKPHONE,3),') ', SUBSTRING(NAD1.T50_WORKPHONE,4,3), '-', RIGHT(NAD1.T50_WORKPHONE,4)) end as "Adjuster Phone"
    ,LTRIM(RTRIM(ISNULL(NAD1.E_MAIL_ADDRESS,''))) as 'EMAIL ADDRESS ADJUSTER'
    ,ISNULL(D74_CLAIM_TYPE,'') as 'Claim_Type'
    ,ISNULL(CC.R30_IN_HOUSE_ADJ,'') as 'Adjuster Code'
    
    INTO ##TEMP_ADD_PRIMIS
    FROM RAW_PRIMIS.PRIMIS2.CCOMMON CC
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.CRESERVE CR
    ON CR.B69_CLAIM_OCCUR = CC.B69_CLAIM_OCCUR
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.PCOMMON PC
    ON PC.A00_PNUM = CC.A00_PNUM
    AND PC.A06_EDITION = CC.A06_EDITION
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.NAME_ADDRESS NA
    ON NA.E04_ORIGNUM COLLATE Latin1_General_CS_AS = CR.E04_CLAIMANT_NUM COLLATE Latin1_General_CS_AS
    AND NA.E04_NEXT IS NULL
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.ACOMMON AC
    ON AC.A04_ANUM COLLATE Latin1_General_CS_AS= PC.A04_ANUM COLLATE Latin1_General_CS_AS
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.NAME_ADDRESS NAD
    ON NAD.E04_ORIGNUM COLLATE Latin1_General_CS_AS  = AC.E04_ANAMNUM COLLATE Latin1_General_CS_AS 
    AND NAD.E04_NEXT IS NULL
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.NAME_ADDRESS NAD1
    ON NAD1.E04_ORIGNUM COLLATE Latin1_General_CS_AS = CASE WHEN NAD1.B27_NAME1 = 'Record Only' THEN CC.R28_SUPERVISOR COLLATE Latin1_General_CS_AS ELSE CC.R30_IN_HOUSE_ADJ COLLATE Latin1_General_CS_AS END
    AND NAD1.E04_NEXT IS NULL
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.NAME_ADDRESS NAD2
    ON NAD2.E04_ORIGNUM COLLATE Latin1_General_CS_AS = CC.E04_INSURED_NUM COLLATE Latin1_General_CS_AS
    AND NAD2.E04_NEXT IS NULL
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.COMPANY CO
    ON CO.A01_COMPANY =CC.A01_COMPANY
    
    LEFT JOIN RAW_PRIMIS.PRIMIS2.NAME_ADDRESS NADC
    ON NADC.E04_ORIGNUM = CO.E04_NAMENUM
    AND NADC.E04_NEXT IS NULL
    
    --LEFT JOIN SANDBOX_CLAIMS_OPS.data.['Directory by Organization'] DBO
    --ON DBO.[Worker] COLLATE SQL_Latin1_General_CP1_CI_AS= ISNULL(ltrim(rtrim(replace(replace(NAD1.B27_NAME1,'(1st)',''),'(3rd)',''))),'')
    
    WHERE AC.J03_PROD_NUM IS NULL
    and NAD1.B27_NAME1 NOT IN  ('To Be Assigned','Void')
    and CR.E87_STATUS <> '4'
    and CC.B69_CLAIM_OCCUR in (
    """ + Primisclaims + """
    )"""
    print('Primis query created!')
    return Primisquery
                         


                              
def runqueryandexcel(source, query, finalquery):
    #Create Connection
    print('Creating server connection...')
    DataLakeserver = '[SERVER_ADDRESS]'
    driver = '{SQL Server}'    # Driver you need to connect to the database
    port = '1433'
    DataLakecnn = pyodbc.connect('DRIVER='+driver+';PORT='+port+';SERVER='+DataLakeserver)
    DataLakecnn.autocommit=True
    
    cursor=DataLakecnn.cursor()
    print("Executing "+source+" query...")
    cursor.execute(query)
    print("Storing "+source+" query results as pd.DF...")
    output = pd.read_sql_query(finalquery, DataLakecnn)
    print(source + " results stored!")                               

    rngoutput = output.values.tolist()
    print('Sheet added!\nRetrieving TEMPLATE workbook')
    wb = load_workbook(r"\\PATH\TO\INPUT\Ack_Template.xlsx")
    print("TEMPLATE workbook retrieved! \nPasting "+source+" output onto Excel sheet..")
    ws=wb.get_sheet_by_name('Extract')
    for row_num, row in enumerate(rngoutput):
        for col_num,val in enumerate(row):
            ws.cell(row=row_num+2,column=col_num+1).value=val #python is zero-indexed, openpyxl is 1-indexed
    print(source+" output pasted onto sheet!")                    

    print('Sheet name updated! \nSaving workbook...')
    wb.save(r"//PATH/TO/OUTPUT/{thedate:%Y}/{thedate:%B}/{today} {source} Ack Letter Extract.xlsx".format(thedate=datetime.date.today(),today=gettoday(False),source=source))
    print('Workbook Saved!')    


### SEND EMAIL


#Set up users for email
#email_user = creds['c1']
#email_pwd = creds['c2']
#Create Module
def mail(to, subject, body, files, username, password, cc=None):
    print("Assigning Sender, Recipient(s), and Subject of email...")
    #Parameters/arguments
    
    msg = MIMEMultipart()
    msg['From'] = username
    msg['To'] = ", ".join(to)
    msg['CC'] = ", ".join(cc)
    bcc=['USER@COMPANY.com',username]
    msg['Subject'] = subject
    print('Adding body of message...')
    msg.attach(MIMEText(body))
    print('Formatting attachments...')
    #get all the attachments
    attachflag = False
    for file in files:
        try: 
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(open(file, 'rb').read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(file))
            msg.attach(part)
            attachflag = True
        except FileNotFoundError:
            pass
   
    if attachflag == False:
        return
    print('Setting up server...')   
    mailServer = smtplib.SMTP('outlook.markelcorp.com', 587)
    mailServer.ehlo()
    mailServer.starttls()
    mailServer.ehlo()
    mailServer.login(username, password)
    mailServer.sendmail(username, to+cc+bcc, msg.as_string())
    # Should be mailServer.quit(), but that crashes...
    mailServer.close()
    print('Email Sent!')

def Mbox(title, text):
    return ctypes.windll.user32.MessageBoxW(0, text, title, 0)
   
#PRIMIS ONLY
#files = [r'\\PATH\TO\FILES\{0} Primis Ack Letter Extract.xlsx'.format(today2)]
#ICON ONLY
#files = [r'\\PATH\TO\FILES\{0} ICON Ack Letter Extract.xlsx'.format(today2)]

if __name__ == '__main__':
    parser = argparse.ArgumentParser(sys.argv)
    parser.add_argument('username', type=str, help="email address")
    parser.add_argument('password', type=str, help = "VDI password")        
    args = parser.parse_args()
    # returns object args
    # args.username : username
    # args.password : password
    main(args.username, args.password)














