import json, pyodbc, pandas as pd, datetime, argparse, sys, os, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
#from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook

def main(email_user, email_pwd, datestring, sendmail:bool=False):
    schedule = get_json()
    start_date, end_date = get_open_and_closed_dates(schedule)
#    datestring = '03/2020'
    d_string = date_format(datestring)
    start_date, end_date = get_open_and_closed_dates(schedule, dates_for=d_string)
    create_folder(end_date)
    cnn = pyodbc.connect('DRIVER={SQL Server}; PORT=1433; SERVER=VA1-PCORSQL191,21612')
    df = run_queries(cnn, end_date)
    excel_save(df, end_date)
    if sendmail:
        send_email(email_user, email_pwd, end_date)
  
def get_json():
    with open(r'\\MKLFILE\CLAIMS\corpfs06-filedrop\ClaimsReporting\Projects\Monthly_Cognos_Reports\Files\schedule.json') as s:
        schedule = json.loads(s.read())
    return schedule


def get_open_and_closed_dates(schedule:dict, dates_for:datetime.date=datetime.date.today()) -> tuple:
    """
    this function finds the start date and end date of the financial month, defaulted to today's month.
    """
    month = str(dates_for.month)
    year = str(dates_for.year)
    start_date = datetime.datetime.strptime(schedule[year][month]['start_date'], '%m/%d/%Y')
    end_date = datetime.datetime.strptime(schedule[year][month]['end_date'], '%m/%d/%Y')
    return start_date, end_date

def create_folder(end_date):
	"""
	this function will look to see if the path, MKLFILE\CLAIMS\corpfs06-filedrop\ClaimsReporting\Monthly Reporting\{end_date:%Y}\{end_date:%m%Y}, exists.  If it doesn't, it will create the folder.  Else, nothing.
	"""
	
	newpathmonth = r"\\MKLFILE\CLAIMS\corpfs06-filedrop\ClaimsReporting\Ad Hoc Reporting\{end_date:%Y}\Jagady Blue - Spec Med & Med Mal".format(end_date=end_date)
	if not os.path.exists(newpathmonth):
		os.makedirs(newpathmonth)
        
def run_queries(cnn, end_date):
    # CREATE CURSOR  
    cursor=cnn.cursor()
    print('Connection established! \nRunning #temp_loss temp table query...')
    # RUN THE QUERIES FOR THE TEMP TABLES
    cursor.execute(temp_loss(end_date))
    print('#temp_loss temp table created! \nRunning #temp_max_reserve temp table query...')
    cursor.execute(temp_max_reserve())
    print('#temp_max_reserve temp table created! \nRunning #temp_reserve temp table query...')
    cursor.execute(temp_reserve(end_date))
    print('#temp_reserve temp table created! \nRunning #temp_detail temp table query...')
    cursor.execute(temp_detail())  
    print('#temp_detail temp table created! \nRunning final_query...')
    # CREATE THE FINAL DATAFRAME
    DF = pd.read_sql_query(final_query(), cnn)
    print('final DF created!')                              
    return DF

def excel_save(DF, end_date):
    rngDF = DF.values.tolist()
    print('Retrieving TEMPLATE workbook')
    wb = load_workbook(r'\\Mklfile\claims\corpfs06-filedrop\ClaimsReporting\Projects\Monthly_Cognos_Reports\Templates\TEMPLATE_MedMal_SpecMed.xlsx')
    print('TEMPLATE workbook retrieved! \nPasting muploadDF onto Excel sheet..')
    ws=wb.get_sheet_by_name('Sheet1')
    for row_num, row in enumerate(rngDF):
        for col_num,val in enumerate(row):
            ws.cell(row=row_num+2,column=col_num+1).value=val #python is zero-indexed, openpyxl is 1-indexed
    print('MedMalDF pasted onto sheet!')   
    print('Sheet name updated! \nSaving workbook...')
    wb.save(r"\\MKLFILE\CLAIMS\corpfs06-filedrop\ClaimsReporting\Ad Hoc Reporting\{end_date:%Y}\Jagady Blue - Spec Med & Med Mal\MedMal_SpecMed_{end_date:%b %Y}.xlsx".format(end_date = end_date))
    print('Workbook Saved!')


def temp_loss(end_date):
    return """IF OBJECT_ID('tempdb..##TEMP_LOSS') IS NOT NULL DROP TABLE ##TEMP_LOSS
                SELECT
                    --CL.[Claim_Feature_Id]
                   distinct cf.Claim_Feature_Count_Key
                   --,cf.Claim_feature_number
                   ,CONCAT(cf.Claim_Folder_Number_Text,'||',cf.Claim_Feature_Number) 'Claim_Feature_Number'
                   ,SUM([Cumulative_Paid_Loss_Net_of_Recoveries_Amount]) AS 'Losses'
                   ,SUM([Cumulative_Paid_ALAE_Net_of_Recoveries_Amount]) AS 'ALAE'
                INTO ##TEMP_LOSS
                FROM [CLAIMS].[dbo].[ADM_CLAIM_CUMULATIVE] CL
                LEFT JOIN CLAIMS.dbo.[ADM_CLAIM_FEATURES] CF
                ON CF.[Claim_Feature_ID] = CL.[Claim_Feature_ID]
                WHERE
                   {end_date:%Y%m} BETWEEN [Effective_From_Accounting_Period_Id] AND [Effective_To_Accounting_Period_ID]
                    AND CL.Direct_Assumed_Ceded_Code <> 'C'
                    AND [Product_Line_Name] = 'Healthcare Risk Solutions'
                    --AND [Product_Line_Name] = 'Medical'
                    AND CF.Suppress_From_Count_Indicator = 'N'
                GROUP BY
                   --CL.[Claim_Feature_ID]
                   CF.Claim_Feature_Count_Key
                  ,CF.Claim_Feature_Number
                  ,cf.Claim_Folder_Number_Text""".format(end_date = end_date)

def temp_max_reserve():
    return """IF OBJECT_ID('tempdb..##TEMP_MAX_RESERVE') IS NOT NULL DROP TABLE ##TEMP_MAX_RESERVE
                    SELECT
                        DISTINCT CF.Claim_Feature_Count_Key
                       ,CONCAT(CF.Claim_Folder_Number_Text,'||',CF.Claim_Feature_Number) 'Claim_Feature_Number'
                       ,MAX([Cumulative_Case_Outstanding_Loss_Reserve_Amount]) AS 'Max Loss Reserves'
                       ,MAX([Cumulative_Case_Outstanding_ALAE_Reserve_Amount]) AS 'Max ALAE Reserves'
                    INTO ##TEMP_MAX_RESERVE
                    FROM [CLAIMS].[dbo].[ADM_CLAIM_CUMULATIVE] CL
                    LEFT JOIN CLAIMS.dbo.[ADM_CLAIM_FEATURES] CF
                    ON CF.[Claim_Feature_ID] = CL.[Claim_Feature_ID]
                    WHERE
                       CL.[Claim_Feature_ID] IS NOT NULL
                        AND CL.Direct_Assumed_Ceded_Code <> 'C'
                        AND [Product_Line_Name] = 'Healthcare Risk Solutions'
                        --AND [Product_Line_Name] = 'Medical'
                        AND CF.Suppress_From_Count_Indicator = 'N'
                    GROUP BY
                       --CL.[Claim_Feature_ID]
                       CF.Claim_Feature_Count_Key
                      ,CF.Claim_feature_number
                      ,CF.Claim_Folder_Number_Text"""
                      
def temp_reserve(end_date):
    return """IF OBJECT_ID('tempdb..##TEMP_RESERVE') IS NOT NULL DROP TABLE ##TEMP_RESERVE
                    SELECT
                        --CL.[Claim_Feature_ID]
                       DISTINCT CF.Claim_Feature_Count_Key
                       ,CONCAT(CF.CLAIM_FOLDER_NUMBER_TEXT,'||', CF.Claim_feature_number) 'Claim_Feature_Number'
                       ,SUM([Cumulative_Case_Outstanding_Loss_Reserve_Amount]) AS 'Loss Reserves'
                       ,SUM([Cumulative_Case_Outstanding_ALAE_Reserve_Amount]) AS 'ALAE Reserves'
                    INTO ##TEMP_RESERVE
                    FROM [CLAIMS].[dbo].[ADM_CLAIM_CUMULATIVE] CL
                    LEFT JOIN CLAIMS.dbo.[ADM_CLAIM_FEATURES] CF
                    ON CF.[Claim_Feature_ID] = CL.[Claim_Feature_ID]
                    WHERE
                       {end_date:%Y%m} BETWEEN [Effective_From_Accounting_Period_Id] AND [Effective_To_Accounting_Period_ID]
                        AND CL.Direct_Assumed_Ceded_Code <> 'C'
                        AND [Product_Line_Name] = 'Healthcare Risk Solutions'
                        --AND [Product_Line_Name] = 'Medical'
                        AND CF.Suppress_From_Count_Indicator = 'N'
                    GROUP BY
                       --CL.[Claim_Feature_ID]
                       CF.Claim_Feature_Count_Key
                      ,CF.Claim_feature_number
                      ,CF.Claim_Folder_Number_Text""".format(end_date = end_date)

def temp_detail():
    return """IF OBJECT_ID('tempdb..##TEMP_DETAIL') IS NOT NULL DROP TABLE ##TEMP_DETAIL
                    SELECT
                       -- CF.[claim_folder_number] AS 'Claim Number'
                       DISTINCT CF.[Claim_Feature_Count_Key] AS 'Claim Feature Key'
                       --,CF.[Claim_Feature_Number] as 'Claim Feature Number'
                       ,CONCAT(CF.CLAIM_FOLDER_NUMBER_TEXT,'||', CF.Claim_feature_number) 'Claim_Feature_Number'
                       ,CF.[Policy_Number] AS 'Policy Number'
                       ,CF.[Policy_Version_Number] AS 'Policy Version'
                       ,CONVERT(VARCHAR(10), CF.[Claim_Feature_Open_Date], 101) AS 'Date Open'
                       ,CONVERT(VARCHAR(10), CF.[Claim_Feature_Closed_Date], 101) AS 'Date Closed'
                       ,ISNULL(R.[Loss Reserves], 0) AS 'Loss Reserves'
                       ,ISNULL(L.Losses, 0) AS 'Losses Paid'
                       ,ISNULL(R.[ALAE Reserves], 0) AS 'ALAE Reserves'
                       ,ISNULL(L.ALAE, 0) AS 'ALAE Paid'
                       ,ISNULL(R.[Loss Reserves] + L.Losses + R.[ALAE Reserves] + L.ALAE, 0) AS 'Total Incurred'
                       ,ISNULL(MR.[Max Loss Reserves], 0) AS 'Max Loss Reserves'
                       ,ISNULL(MR.[Max ALAE Reserves], 0) AS 'Max ALAE Reserves'
                       ,CF.[Claimant_Name] AS 'Claimant'
                       ,CF.[Claim_Examiner_Name] AS 'Examiner'
                       ,CF.[Product_Line_Name] AS 'Product Line 1'
                       ,CF.[Product_Line2_Name] AS 'Product Line 2'
                       ,CF.[ISO_Country_Subdivision_2Digit_Name] AS 'Loss State'
                       ,CF.[Person_Name_Preferred] AS 'Underwriter'
                       ,CF.[Division_Name] AS 'UW Division'
                       ,CF.[Subdivision_Name] AS 'Subdivision'
                       ,CONVERT(VARCHAR(10), CF.[Reported_Date], 101) AS 'Date Reported'
                       ,CONVERT(VARCHAR(10), CF.[Date_of_Loss_Date], 101) AS 'Date of Loss'
                       ,CONVERT(VARCHAR(10), CF.[Claim_Feature_Re_Open_Date], 101) AS 'Date Re-Open'
                       ,CONVERT(VARCHAR(10), CF.[Claims_Made_Date], 101) AS 'Date Claims Made'
                       ,CF.[Insured_Name] AS 'Insured'
                       ,CF.[Parent_External_Source_Name] AS 'Source'
                    INTO ##TEMP_DETAIL
                    FROM [CLAIMS].[dbo].[ADM_CLAIM_FEATURES] CF
                    LEFT JOIN ##TEMP_LOSS L
                       ON L.[Claim_Feature_Number] = CONCAT(CF.CLAIM_FOLDER_NUMBER_TEXT,'||', CF.Claim_feature_number)
                    LEFT JOIN ##TEMP_RESERVE R
                       ON R.[Claim_Feature_Number] = CONCAT(CF.CLAIM_FOLDER_NUMBER_TEXT,'||', CF.Claim_feature_number)
                    LEFT JOIN ##TEMP_MAX_RESERVE MR
                       ON MR.[Claim_Feature_Number] = CONCAT(CF.CLAIM_FOLDER_NUMBER_TEXT,'||', CF.Claim_feature_number)
                    WHERE
                       CF.[Claim_Feature_ID] IS NOT NULL
                          AND CF.[Claim_Folder_sub_Status_code] <> 'V'
                        AND [Product_Line_Name] = 'Healthcare Risk Solutions'
                        --AND [Product_Line_Name] = 'Medical'
                        AND CF.Suppress_From_Count_Indicator = 'N'
                    ORDER BY
                    ISNULL(R.[Loss Reserves] + L.Losses + R.[ALAE Reserves] + L.ALAE, 0) DESC"""

def final_query():
    return """SELECT 
                --CF.[Claim Feature Key]
                  REPLACE(CF.[Claim_Feature_Number],'||','-') 'Claim Feature Number'
                  --,CF.[Claim Feature Number1]
                  ,SS.[Claim_Folder_Substatus_Name]
                  ,SUB.[Claim_Feature_Substatus_Name]
                  ,DIV.[Rating_Class_Markel_Subdivision_Description_Text]
                  ,CC.Accident_Year_Number
                  ,FF.Legacy_Claim_Folder_Type_Code
                  ,CASE WHEN CF.Source = 'PRIMIS' then CC.Legacy_Cause_of_Loss_Description_Text ELSE  CC.Legacy_Type_of_Loss_Description_Text END AS 'Loss Description'
                  ,CF.[Policy Number]
                  ,CF.[Policy Version]
                  ,CF.[Date Open]
                  ,CF.[Date Closed]
                  ,CF.[Loss Reserves]
                  ,CF.[Losses Paid]
                  ,CF.[ALAE Reserves]
                  ,CF.[ALAE Paid]
                  ,CF.[Total Incurred]
                  ,CF.[Max Loss Reserves]
                  ,CF.[Max ALAE Reserves]
                  ,CF.[Claimant]
                  ,CF.[Examiner]
                	 ,CF.[UW Division]
                  ,CF.[Subdivision]
                  ,CF.[Product Line 1]
                  ,CF.[Product Line 2]
                  ,CF.[Loss State]
                  ,CF.[Underwriter]
                  ,CF.[Date Reported]
                  ,CF.[Date of Loss]
                  ,CF.[Date Re-Open]
                  ,CF.[Date Claims Made]
                  ,CF.[Insured]
                  ,CF.[Source]
              FROM ##TEMP_DETAIL CF
              LEFT JOIN [ADMPROD].ADM.[dbo].[Dim_Claim_Feature_Base_Extended] CC
              ON CC.External_Reference_Text = CF.[Claim_Feature_NUmber]
              LEFT JOIN [ADMPROD].ADM.[dbo].[Dim_Claim_Feature_Substatus] SUB
              on SUB.[Claim_Feature_Substatus_Id] = CC.[Claim_Feature_Substatus_Id]
              LEFT JOIN [ADMPROD].ADM.[dbo].[Dim_Claim_Folder_Base_Extended] FF
              ON FF.Claim_Folder_Id = CC.Claim_Folder_Claim_Id
              AND FF.External_Source_Code_Id = CC.Claim_Folder_Claim_External_Source_Code_Id
              LEFT JOIN [ADMPROD].ADM.[dbo].[Dim_Claim_Folder_Substatus] SS
              ON SS.[Claim_Folder_Substatus_Id] = FF.[Claim_Folder_Substatus_Id]
              LEFT JOIN [ADMPROD].ADM.[dbo].[Dim_Coverage_Component_Base_Extended] X
              ON X.[COVC_Agreement_Version_Id] = CC.[COVC_Agreement_Version_Id]
              AND X.COVC_Agreement_External_Source_Code_Id = CC.COVC_Agreement_Version_External_Source_Code_Id
              LEFT JOIN  [ADMPROD].ADM.[dbo].[Dim_Rating_Class_Markel_Subdivision] DIV
              ON DIV.[Rating_Class_Markel_Subdivision_Id]  = X.[Rating_Class_Markel_Subdivision_Id]
              WHERE CC.Suppress_From_Count_Indicator = 'N'"""

def date_format(datestring):
    x = datestring
    x=datetime.datetime.strptime(x,'%m/%Y')
    return x


#Create Module
def mail(email_user, email_pwd, to, subject, text, attach, cc=None):
   print("Assigning Sender, Recipient(s), and Subject of email...")
   msg = MIMEMultipart()
   msg['From'] = email_user
   msg['To'] = ", ".join(to)
   msg['CC'] = ", ".join(cc)
   bcc=['whan@markelcorp.com']
   msg['Subject'] = subject
   print('Adding body of message...')
   msg.attach(MIMEText(text))
   print('Formatting attachments...')
   #get all the attachments
   for file in attach:
      part = MIMEBase('application', 'octet-stream')
      part.set_payload(open(file, 'rb').read())
      encoders.encode_base64(part)
      part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(file))
      msg.attach(part)
   print('Setting up server...')   
   mailServer = smtplib.SMTP('outlook.markelcorp.com', 587)
   mailServer.ehlo()
   mailServer.starttls()
   mailServer.ehlo()
   mailServer.login(email_user, email_pwd)
   mailServer.sendmail(email_user, to+cc+bcc, msg.as_string())
   # Should be mailServer.quit(), but that crashes...
   mailServer.close()
   print('Email Sent!')


def send_email(email_user, email_pwd, end_date):
    #Parameters/arguments
    recipients = ['jblue@markelcorp.com', 'pmoylan@markelcorp.com']
    cc = ['rkincaid@markelcorp.com']
    subject = "{end_date:%b %Y} Spec Med Med Mal Report".format(end_date=end_date)
    body = "Hello,\n\nAttached you will find the Spec Med and Med Mal report for {end_date:%b %Y} MEFC.  If you have any questions, please feel free to contact me.\n\nThanks,\nWill Han".format(end_date = end_date)
    #Set up crap for the attachments
    #FOR ICON AND PRIMIS
    files = [r"\\MKLFILE\CLAIMS\corpfs06-filedrop\ClaimsReporting\Ad Hoc Reporting\{end_date:%Y}\Jagady Blue - Spec Med & Med Mal\MedMal_SpecMed_{end_date:%b %Y}.xlsx".format(end_date = end_date)]
    mail(email_user, email_pwd, recipients, subject, body, files, cc)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(sys.argv)
    parser.add_argument('-e', type = str, help = 'your outlook email address')
    parser.add_argument('-p', type = str, help = 'your outlook email password (same as your VDI)')
    parser.add_argument('-d', type = str, help = 'date string')
    parser.add_argument('--creds', type = str, help = 'the path to your credentials file') # -- is optional
    args = parser.args()
    if (args.e and args.p and args.d):
        main(args.e, args.p, args.d)
    elif (args.creds):
        with open(args.creds) as c:
            creds = json.loads(c.read())
        email = creds['c1']
        pw = creds['c2']
        main(email, pw)
    else:
        raise Exception('Input either your email address and password OR the path to your config file')
            
        