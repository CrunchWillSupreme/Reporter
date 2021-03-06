"""
			ADM Monthlies for CLI
			Created 10/11/18
			Author: Will Han
"""
import pyodbc, pandas as pd, datetime as dt, os, smtplib, json, time, argparse, sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from smtplib import SMTPSenderRefused
from pandas.io.sql import DatabaseError
import cx_Oracle


def main(email_user:str, email_pwd:str, datestring, sendmail:bool=False):
    """
    This is the Cognos ADM monthly reporting weapon.  
    	
    Parameters:
    email_user: your outlook server email address. 
    email_pwd: your outlook server email password (same password as your VDI)
    	
    Process:
    1.  Load in config files (CognosADM and schedule)
    2.  Retrieve current open and close date for the system month.
    3.  Format the config file data to substitute the correct dates.
    - For each report in the config file:
    4.  Create the report in excel using the SQL server driver pyodbc.
    5.  Email the report to the corresponding recipient(s).
    """
    CognosADM, CognosICON, schedule = get_json()
    d_string = date_format(datestring)
    start_date, end_date = get_open_and_closed_dates(schedule, dates_for=d_string)
    format_config(CognosADM, start_date, end_date)
    format_config(CognosICON, start_date, end_date)
    create_folder(end_date)
#    cnn = pyodbc.connect('DRIVER={SQL Server}; PORT=1433; SERVER=[SQL_SERVER_ADDRESS]')
#    print("Oracle Connection Established!")
	
    rpts = []
    for ADMreport, ADMdata in CognosADM.items():
        print('running '+ADMreport)
        report = Report(ADMreport, schedule, start_date, end_date)
        rpts.append(report)
    for ICONreport, ICONdata in CognosICON.items():
        print('running '+ICONreport)
        report = ICONReport(ICONreport, schedule)
        rpts.append(report)
    for instance in rpts:
        instance.run_report()
 
    if sendmail:
        for occur in rpts:
            flag = False
            while not flag:
                try:
                    occur.send_email(email_user, email_pwd)
                    flag = True
                    print('Email Sent!')
                except SMTPSenderRefused:
                    time.sleep(5)

def single_report(name:str, email_user:str, email_pwd:str, datestring, sendmail:bool=False):
    CognosADM, CognosICON, schedule = get_json()
    d_string = date_format(datestring)
    start_date, end_date = get_open_and_closed_dates(schedule, dates_for=d_string)
    format_config(CognosADM, start_date, end_date)
    format_config(CognosICON, start_date, end_date)
    create_folder(end_date)
    rpts = []
    if name in CognosADM:
        report = Report(CognosADM[name], schedule,start_date, end_date)
        rpts.append(report)
    elif name in CognosICON:
        report = ICONReport(CognosICON[name], schedule, start_date, end_date)
        rpts.append(report)
    else:
        raise Exception("Could not find "+name+" in either config dicts.")
    for instance in rpts:
        instance.run_report()
    if sendmail:
        for occur in rpts:
            flag = False
            while not flag:
                try:
                    occur.send_email(email_user, email_pwd)
                    flag = True
                    print('Email Sent!')
                except SMTPSenderRefused:
                    time.sleep(5)
   

class Report(object):
    """
    a class that creates instances of the base report model
    """
    def __init__(self,config:dict, schedule:dict, start_date:dt.datetime, end_date:dt.datetime):
        self.sql_file = config.get('sql',None)
        self.template_file = config['template']
        self.save_as = config['save_as']
        self.recipients = config.get('recipients',None)
        self.subject = config.get('subject',None)
        self.body =config.get('body',None)
        self.attachs = config.get('attachs',None)
        self.ccopy = config.get('ccopy',None)
        self.sum_start = config.get('sum_start',None)
        self.sum_end = config.get('sum_end',None)
        self.start_date = start_date
        self.end_date = end_date
        self.temp_tables = config.get('temp_tables',None)
    
    def get_conn(self):
        print('Getting Connection...')
        return pyodbc.connect('DRIVER={SQL Server}; PORT=1433; SERVER=[SQL_SERVER_ADDRESS]')
        
    def run_report(self):
        query = self.get_query()
        cnn = self.get_conn()
        self.load_temp_tables(cnn)
        df = self.run_query(query,cnn)
        formatted = self.format_df(df)
        self.savewb(formatted)

    def get_query(self):
        print('Getting Query...')
        with open(r'\\PATH\TO\SQL\FILES\{0}'.format(self.sql_file)) as f:
            query=f.read()
        return query
    
    def load_temp_tables(self, cnn):
        print('Checking for Temp Tables...')
        if not self.temp_tables:
            return
        for table in self.temp_tables:
            print('Opening Temp tables...')
            with open(table) as t:
                temp = t.read()
            cursor=cnn.cursor()
            cursor.execute(temp)
        
    def run_query(self,query,cnn):
        print('Running Query(ies)...')
        try:
            df=pd.read_sql(query, cnn, params=(self.start_date.strftime("%x"), self.end_date.strftime("%x")))
        except DatabaseError: 
            df=pd.read_sql(query, cnn) ##FOR QUERIES THAT DON'T HAVE DATE PARAMETERS
        return df
        print(self.SQLfile[:-4]+" Query Executed!")

        
    def format_df(self,df):
        print("Adding summary row...")
        formatted = df.copy()
        formatted.loc['Summary'] = pd.Series(formatted.iloc[:,self.sum_start:self.sum_end].sum())
        formatted.iloc[-1,0] = 'Summary'
        print("Summary row added!")
        return formatted
        
        
    def savewb(self, formatted):
        print('Beginning savewb method...')
        rngoutput = formatted.values.tolist()
        print('Sheet added!\nRetrieving TEMPLATE workbook')
        wb = load_workbook(r"\\PATH\TO\TEMPLATES\{0}".format(self.template_file))
        print("TEMPLATE workbook retrieved! \nPasting df output onto Excel sheet..")
        ws=wb.get_sheet_by_name('Page1')
        for row_num, row in enumerate(rngoutput):
            for col_num,val in enumerate(row):
                ws.cell(row=row_num+5,column=col_num+1).value=val #python is zero-indexed, openpyxl is 1-indexed
        print(self.sql_file[:-4]+" output pasted onto sheet!\nSaving workbook...")                    
        wb.save(r"\\PATH\TO\OUTPUT\{end_date:%m%Y}\{save_as}.xlsx".format(end_date = self.end_date, save_as=self.save_as))
        print(self.sql_file[:-4]+' Workbook Saved!')
    
    def send_email(self, email_user, email_pwd):
        print("Assigning Sender, Recipient(s), and Subject of email...")
        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = ", ".join(self.recipients)
        msg['CC'] = ", ".join(self.ccopy)
        bcc=['whan@COMPANY.com']
        msg['Subject'] = self.subject
        print('Adding body of message...')
        msg.attach(MIMEText(self.body))
        print('Formatting attachments...')
        #get all the attachments
        for attc in self.attachs:
           print(attc) 
           part = MIMEBase('application', 'octet-stream')
           part.set_payload(open(attc, 'rb').read())
           encoders.encode_base64(part)
           part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(attc))
           msg.attach(part)
        print('Setting up server...')   
        server = self.get_server(email_user, email_pwd)
        server.sendmail(email_user, self.recipients+self.ccopy+bcc, msg.as_string())
        
    def get_server(self, email_user, email_pwd):
        mailServer = smtplib.SMTP('outlook.markelcorp.com', 587)
        mailServer.ehlo()
        mailServer.starttls()
        mailServer.ehlo()
        mailServer.login(email_user, email_pwd)
        return mailServer


# Subclass for ICON connection
class ICONReport(Report):
    def get_conn(self):
        connection = cx_Oracle.connect('cog{end_date:%y%m}/cog{end_date:%y%m}@m[PORT:ADDRESS]/[PARAMETER]'.format(end_date=self.end_date))
        return connection    
        
def get_json():
    """
    this function gets the config files, CognosADMfiles.json and schedule.json, and loads them in as a dictionary.
    """	
    with open(r'\\PATH\TO\CONFIG\CognosADMFiles.json') as f:
        CognosADM = json.loads(f.read())
    with open(r'\\PATH\TO\CONFIG\CognosIconFiles.json') as f:
        CognosICON = json.loads(f.read()) 
    with open(r'\\PATH\TO\CONFIG\schedule.json') as s:
        schedule = json.loads(s.read())
    return CognosADM, CognosICON, schedule

def get_open_and_closed_dates(schedule:dict, dates_for:dt.date=dt.date.today()) -> tuple:
    """
    this function finds the start date and end date of the financial month, defaulted to today's month.
    """
    	
    month = str(dates_for.month)
    year = str(dates_for.year)
    start_date = dt.datetime.strptime(schedule[year][month]['start_date'], '%m/%d/%Y')
    end_date = dt.datetime.strptime(schedule[year][month]['end_date'], '%m/%d/%Y')
    return start_date, end_date
	
def format_config(CognosDict:dict, start_date, end_date):
	"""
	this function formats the config file data to replace the {formats} in the file with the values of the variables (start_date, end_date, etc..)
	"""
	
	for report, data in CognosDict.items():
		formats =  {
				'start_date':start_date,
				'end_date':end_date,
#				'today':today,
				'save_as':data['save_as']
				}
		try:
			for i, file in enumerate(data['attachs']):
				data['attachs'][i] = file.format(**formats)
		except:
			print('The report, '+report+', does not have any attachments.')
		for key, field in data.items():
			if not isinstance(field, str):
				continue
			data[key] = field.format(**formats)
	return

def create_folder(end_date):
	"""
	this function will look to see if the path, PATH\TO\REPORTING\{end_date:%Y}\{end_date:%m%Y}, exists.  If it doesn't, it will create the folder.  Else, nothing.
	"""
	
	newpathmonth = r'\\PATH\TO\REPORTING\{end_date:%Y}\test\{end_date:%m%Y}'.format(end_date=end_date)
	if not os.path.exists(newpathmonth):
		os.makedirs(newpathmonth)

def date_format(datestring):
    x = datestring
    x=dt.datetime.strptime(x,'%m/%Y')
    return x
    
   

if __name__ == '__main__':
    parser = argparse.ArgumentParser(sys.argv)
    parser.add_argument('-e', type = str, help = 'email')
    parser.add_argument('-p', type = str, help = 'password')
    parser.add_argument('-d', type = str, help = 'start date (month/year)')
    parser.add_argument('--credfile', type = str, help = 'credentials file.json') # -- is optional
    args = parser.parse_args()
    if (args.e and args.p and args.d):
        main(args.e, args.p, args.d)
#    elif (args.credfile):
#    with open(args.credfile) as c:
#    creds = json.loads(c.read())
#    email = creds['c1']
#    pw = creds['c2']
#    main(email, pw)
    else:
#        raise Exception('Either input your email address and password or the path to your config file (json)')
        raise Exception('Either input your email address and password and the start date')
	
	
	
	
	
	
	
