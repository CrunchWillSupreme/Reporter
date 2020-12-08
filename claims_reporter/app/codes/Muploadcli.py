"""
Matter Upload CLI
"""

import pandas as pd, argparse, sys, pyodbc, datetime, json, os, smtplib
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders


def main(email_user, email_pwd, send_mail:bool=False):
    print('Beginning Matter Upload..')
    d= datetime.date.today()
    month=str(d.strftime('%m'))
    day=str(d.strftime('%d'))
    year= str(d.year)
    today= year + '_' + month + '_' + day
    cnn = pyodbc.connect('DRIVER={SQL Server};PORT=1433;SERVER=[SERVER_ADDRESS]')
    run_temp_tables(cnn)
    DF = run_query(cnn)
    write_excel(DF, year, today)
    if send_mail:
        recipients = ['USER@EMAIL.com']
        cc = ['USER@EMAIL.com']
        subject = "LeX Matter Upload " + today
        body = "Hi USER,\n\nAttached you will find this week's LeX Matter Upload.  Please let me know if you have any questions.\n\nThanks,\nWill Han"
        #Set up crap for the attachments
        files = [r"\\PATH\TO\FILE\{year}\{today} LeX Matter Upload.xlsx".format(year = year, today = today)]
        email(email_user, email_pwd, recipients, cc, subject, body, files)
        
def run_temp_tables(cnn):
    print('Reading in temp tables..')
    with open(r'\\PATH\TO\FILE\base.sql') as b:
        base = b.read()
    with open(r'\\PATH\TO\FILE\substit.sql') as s:
        substit = s.read()
    with open(r'\\PATH\TO\FILE\types.sql') as t:
        types = t.read()
    with open(r'\\PATH\TO\FILE\deduct.sql') as d:
        deduct = d.read()
    with open(r'\\PATH\TO\FILE\smash.sql') as m:
        smash = m.read()
    print('Temp tables read! Executing temp table queries..')
    cursor = cnn.cursor()
    cursor.execute(base)
    cursor.execute(substit)
    cursor.execute(types)
    cursor.execute(deduct)
    cursor.execute(smash)
    print('Temp tables executed!')
    
def run_query(cnn):
    print('Running final query..')
    MuploadDF = pd.read_sql_query("""select * from #SMASH """, cnn)
    print('Final query ran!')
    return MuploadDF

def write_excel(df, year, today):
    print('Writing data to Excel..')
    rngMupload = df.values.tolist()
    wb = load_workbook(r'\\PATH\TO\FILE\Matter_Upload_Template.xlsx')
    ws=wb.get_sheet_by_name('template')
    for row_num, row in enumerate(rngMupload):
        for col_num,val in enumerate(row):
            ws.cell(row=row_num+2,column=col_num+1).value=val
    print('Sheet name updated! \nSaving workbook...')
    wb.save(r"\\PATH\TO\FILE\{year}\{today} LeX Matter Upload.xlsx".format(year=year, today = today))
    print('Workbook Saved!')
#x=f"\\PATH\TO\FILE\{year}\{today} LeX Matter Upload.xlsx"

def set_server(email_user, email_pwd):
    print('Setting up server...')   
    mailServer = smtplib.SMTP('outlook.markelcorp.com', 587)
    mailServer.ehlo()
    mailServer.starttls()
    mailServer.ehlo()
    mailServer.login(email_user, email_pwd)
    return mailServer
   
   
def email(email_user, email_pwd, to, cc, subject, text, attach):
    print("Assigning Sender, Recipient(s), and Subject of email...")
    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = ", ".join(to)
    msg['CC'] = ", ".join(cc)
    bcc=['USER@EMAIL.com']
    msg['Subject'] = subject
    print('Adding body of message...')
    msg.attach(MIMEText(text))
    print('Formatting attachments...')
    #get all the attachments
    for file in attach:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(open(file, 'rb').read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(file))
        msg.attach(part)
    server = set_server(email_user, email_pwd)
    server.sendmail(email_user, to+cc+bcc, msg.as_string())
    # Should be mailServer.quit(), but that crashes...
    server.close()
    print('Email Sent!')

if __name__ == '__main__':
    parser = argparse.ArgumentParser(sys.argv)
    parser.add_argument('username', type=str, help="email address")
    parser.add_argument('password', type=str, help = "VDI password")    
    args = parser.parse_args()
    # returns object args
    # args.username : username
    # args.password : password
    main(args.username, args.password)
    
    
